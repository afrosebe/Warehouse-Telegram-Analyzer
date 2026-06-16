#!/usr/bin/env python3
"""
Warehouse Auction Analyzer - Telegram Bot v3
Fixed: Tentative Final, Report Download, Column Mapping
Render.com deployment ready
"""
import os, sys, logging, tempfile
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes
import openpyxl

BOT_TOKEN = os.environ.get('BOT_TOKEN', '')
if not BOT_TOKEN:
    print("BOT_TOKEN not set!")
    sys.exit(1)

logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)
user_data = {}

def sf(v):
    if v is None or v == '' or v == '#N/A' or v == '!REF': return 0
    if isinstance(v, (int, float)): return float(v) if v == v else 0
    try:
        s = str(v).replace(',', '').strip()
        return float(s) if s and s != '-' else 0
    except: return 0

def fmt(v):
    v = round(abs(v), 2)
    if v >= 100000: return f'₹{v/100000:.2f}L'
    if v >= 1000: return f'₹{v/1000:.1f}K'
    return f'₹{v:.2f}'

def find_columns(headers):
    col = {}
    for i, h in enumerate(headers):
        hl = str(h).strip().lower() if h else ''
        if 'warehouse_id' in hl or ('warehouse' in hl and 'id' in hl): col['wh'] = i
        elif hl == 'brand': col['brand'] = i
        elif 'product_title' in hl or ('product' in hl and 'title' in hl): col['title'] = i
        elif hl == 'mrp': col['mrp'] = i
        elif hl == 'available_quantity' or hl == 'quantity': col['qty'] = i
        elif 'manufacturing' in hl or hl == 'mfg_date': col['mfg'] = i
        elif 'expiry' in hl or 'expire' in hl: col['exp'] = i
        elif hl == 'aging' or hl == 'age': col['aging'] = i
        elif hl == 'slab': col['slab'] = i
        elif 'yield' in hl: col['yield'] = i
        elif 'category' in hl or 'super_category' in hl: col['cat'] = i
        elif 'tenative' in hl: col['tenative'] = i
        elif 'tcs value' in hl: col['tcsval'] = i
        elif hl == 'tcs': col['tcs'] = i
        elif 'tentative final' in hl or 'tentative_final' in hl: col['tf'] = i
        elif 'final mrp' in hl or 'final_mrp' in hl: col['fm'] = i
    return col

def analyze_warehouse(rows, headers, target_wh):
    col = find_columns(headers)
    wh_idx = col.get('wh')
    if wh_idx is None: return None
    filtered = [r for r in rows if str(r[wh_idx]).strip() == target_wh]
    if not filtered: return None
    prods = []
    tQ = tTF = tFM = tMRP = 0
    for r in filtered:
        mrp = sf(r[col.get('mrp', 12)])
        qty = sf(r[col.get('qty', 13)])
        tf = sf(r[col.get('tf', 28)])
        fm = sf(r[col.get('fm', 29)])
        aging = sf(r[col.get('aging', 21)])
        days_left = int(aging)
        exp_idx = col.get('exp', 17)
        if exp_idx is not None and r[exp_idx]:
            try:
                ed = r[exp_idx]
                if isinstance(ed, datetime):
                    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                    days_left = (ed - today).days
            except: pass
        if days_left <= 30: pri = '🔴 CRITICAL'
        elif days_left <= 60: pri = '🟡 MEDIUM'
        else: pri = '🟢 SAFE'
        mfg_s = ''
        mfg_idx = col.get('mfg', 16)
        if mfg_idx is not None and r[mfg_idx]:
            try: mfg_s = r[mfg_idx].strftime('%d-%b') if isinstance(r[mfg_idx], datetime) else str(r[mfg_idx])[:10]
            except: pass
        exp_s = ''
        if exp_idx is not None and r[exp_idx]:
try: exp_s = r[exp_idx].strftime('%d-%b') if isinstance(r[exp_idx], datetime) else str(r[exp_idx])[:10]
            except: pass
        prods.append({
            'brand': str(r[col.get('brand', 10)] or ''),
            'title': str(r[col.get('title', 11)] or ''),
            'cat': str(r[col.get('cat', 8)] or ''),
            'mrp': mrp, 'qty': qty, 'mfg': mfg_s, 'exp': exp_s,
            'days': days_left, 'slab': str(r[col.get('slab', 22)] or ''),
            'yield': sf(r[col.get('yield', 24)]),
            'tf': tf, 'fm': fm, 'pri': pri
        })
        tQ += qty; tTF += tf; tFM += fm; tMRP += mrp * qty
    disc = ((tMRP - tTF) / tMRP * 100) if tMRP > 0 else 0
    return {
        'wh': target_wh, 'prods': prods,
        'tQ': int(tQ), 'tTF': round(tTF, 2), 'tFM': round(tFM, 2), 'tMRP': round(tMRP, 2),
        'disc': round(disc, 1)
    }

def format_analysis(a):
    wh = a['wh']
    prods = a['prods']
    if a['disc'] >= 25: verdict = '✅ STRONG BUY — High discount, great deal!'
    elif a['disc'] >= 15: verdict = '🟡 CONSIDER — Moderate discount'
    elif a['disc'] >= 10: verdict = '⚠️ WEAK BUY — Low discount'
    else: verdict = '❌ SKIP — Not enough discount'
    crit = [p for p in prods if 'CRITICAL' in p['pri']]
    med = [p for p in prods if 'MEDIUM' in p['pri']]
    safe_p = [p for p in prods if 'SAFE' in p['pri']]
    cats = {}
    for p in prods:
        c = p['cat'] or 'Unknown'
        cats[c] = cats.get(c, 0) + p['qty']
    msg = f"""📊 *Analysis: {wh}*

📦 Products: {len(prods)}
📊 Total Qty: {a['tQ']:,}
💰 MRP Value: {fmt(a['tMRP'])}
🎯 Tentative Final: {fmt(a['tTF'])}
📋 Final MRP: {fmt(a['tFM'])}
📉 Discount: *{a['disc']:.1f}%*

{verdict}

🔴 Critical (≤30d): {len(crit)} ({fmt(sum(p['tf'] for p in crit))})
🟡 Medium (31-60d): {len(med)} ({fmt(sum(p['tf'] for p in med))})
🟢 Safe (60+d): {len(safe_p)} ({fmt(sum(p['tf'] for p in safe_p))})

📦 Categories:"""
    for c, q in sorted(cats.items(), key=lambda x: -x[1])[:5]:
        pct = (q / a['tQ'] * 100) if a['tQ'] > 0 else 0
        msg += f"\n  • {c}: {int(q):,} ({pct:.0f}%)"
    if crit:
        msg += "\n\n🔴 Top Critical Items:"
        for p in sorted(crit, key=lambda x: x['days'])[:5]:
            msg += f"\n  • {p['brand']} - {p['title'][:30]}"
            msg += f"\n    Qty: {int(p['qty'])} | Days: {p['days']} | TF: {fmt(p['tf'])}"
    msg += "\n\n💡 /report bhejo Excel report download karne ke liye!"
    return msg

def generate_report_excel(a):
    logger.info(f"Generating report for warehouse: {a.get('wh', 'unknown')}")
    out_wb = openpyxl.Workbook()
    ws1 = out_wb.active
    ws1.title = 'Analysis'
    hr = ['S.No','Brand','Product','Category','MRP','Qty','Mfg','Expiry','Days','Slab','Yield','Tenative','TCS','TCS Val','Tentative Final','Final MRP','Priority']
    for ci, hv in enumerate(hr, 1):
        c = ws1.cell(row=1, column=ci, value=hv)
        c.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        c.fill = openpyxl.styles.PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for ri, p in enumerate(a['prods'], 2):
        vals = [ri-1, p['brand'], p['title'], p['cat'], p['mrp'], int(p['qty']),
                p['mfg'], p['exp'], p['days'], p['slab'], p['yield'], 0, 0, 0,
                round(p['tf'],2), round(p['fm'],2), p['pri']]
        for ci, v in enumerate(vals, 1):
            ws1.cell(row=ri, column=ci, value=v)
    tr = len(a['prods']) + 2
    for c in [1, 5, 15, 16]:
        ws1.cell(row=tr, column=c).font = openpyxl.styles.Font(bold=True)
    ws1.cell(row=tr, column=1, value='TOTAL')
    ws1.cell(row=tr, column=5, value=a['tQ'])
    ws1.cell(row=tr, column=15, value=a['tTF'])
    ws1.cell(row=tr, column=16, value=a['tFM'])
    ws2 = out_wb.create_sheet('Decision')
    ws2.append(['PURCHASE DECISION'])
    ws2.append([])
    ws2.append(['Products', len(a['prods'])])
    ws2.append(['Quantity', a['tQ']])
    ws2.append(['MRP Value', f'₹{a["tMRP"]:.0f}'])
    ws2.append(['Tentative Final', f'₹{a["tTF"]:.2f}'])
ws2.append(['Final MRP', f'₹{a["tFM"]:.2f}'])
    ws2.append(['Discount %', f'{a["disc"]:.1f}%'])
    ws2.append([])
    ws2.append(['Verdict', 'STRONG BUY' if a['disc'] >= 25 else 'CONSIDER' if a['disc'] >= 15 else 'WEAK BUY' if a['disc'] >= 10 else 'SKIP'])
    ws2.append([])
    ws2.append(['Priority', 'Products', 'Tentative Final', 'Final MRP'])
    crit = [p for p in a['prods'] if 'CRITICAL' in p['pri']]
    med = [p for p in a['prods'] if 'MEDIUM' in p['pri']]
    safe_p = [p for p in a['prods'] if 'SAFE' in p['pri']]
    for name, grp in [('CRITICAL', crit), ('MEDIUM', med), ('SAFE', safe_p)]:
        ws2.append([name, len(grp), f'₹{sum(p["tf"] for p in grp):.2f}', f'₹{sum(p["fm"] for p in grp):.2f}'])
    ws3 = out_wb.create_sheet('Critical')
    ws3.append(['#','Brand','Product','MRP','Qty','Expiry','Days','Tentative Final','Final MRP'])
    for ri, p in enumerate(sorted(crit, key=lambda x: x['days']), 1):
        ws3.append([ri, p['brand'], p['title'], p['mrp'], int(p['qty']), p['exp'], p['days'], round(p['tf'],2), round(p['fm'],2)])
    import tempfile as tf
    tmp = tf.NamedTemporaryFile(suffix='.xlsx', delete=False)
    out_wb.save(tmp.name)
    tmp.close()
    logger.info(f"Report saved: {tmp.name}")
    return tmp.name

async def start(update: Update, context):
    welcome = """
🤖 *Warehouse Auction Analyzer Bot*

Mera naam hai Warehouse Bot! Main tumhari Excel file analyze karke auction mein kitna discount mil sakta hai, woh bataata hoon.

📌 *Kaise use karein:*
1. Mujhe Excel file bhejo (.xlsx)
2. Warehouse select karo (button dabao)
3. Instant analysis!

📌 *Commands:*
/start - Welcome message
/help - Help
/report - Last analysis ka Excel report download
/all - Saare warehouse ka summary

💡 *Tip:* Seedha Excel file bhejo, main khud warehouse list dunga!
    """
    await update.message.reply_text(welcome, parse_mode='Markdown')

async def help_cmd(update: Update, context):
    help_text = """
📖 *Help*

1. Excel file bhejo (.xlsx format)
2. Main warehouse list dunga with product count
3. Warehouse select karo (button dabao)
4. Analysis ready!

📊 *Analysis mein milega:*
• Tentative Final (₹) — total estimated cost
• Final MRP (₹) — total MRP value
• Discount % — kitna discount mil sakta hai
• 🔴 Critical items (≤30 days expiry)
• 🟡 Medium (31-60 days)
• 🟢 Safe (60+ days)
• Category breakdown
• Downloadable Excel Report
    """
    await update.message.reply_text(help_text, parse_mode='Markdown')

async def handle_document(update: Update, context):
    document = update.message.document
    file_name = document.file_name
    if not file_name.lower().endswith(('.xlsx', '.xls')):
        await update.message.reply_text("❌ Sirf .xlsx ya .xls file bhejo!")
        return
    status_msg = await update.message.reply_text("📥 File download ho rahi hai...")
    try:
        file = await context.bot.get_file(document.file_id)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_name = tmp.name
        await file.download_to_drive(tmp_name)
        await status_msg.edit_text("📊 File analyze ho rahi hai...")
        wb = openpyxl.load_workbook(tmp_name, data_only=True)
        ws = None
        headers = None
        wh_col_idx = None
        for sn in wb.sheetnames:
            sh = wb[sn]
            if sh.max_row < 2: continue
            hdrs = [str(c.value).strip() if c.value else '' for c in sh[1]]
            for i, h in enumerate(hdrs):
                if 'warehouse' in h.lower():
                    ws = sh; headers = hdrs; wh_col_idx = i; break
            if wh_col_idx is not None: break
            if sh.max_row >= 3:
                hdrs = [str(c.value).strip() if c.value else '' for c in sh[2]]
                for i, h in enumerate(hdrs):
                    if 'warehouse' in h.lower():
                        ws = sh; headers = hdrs; wh_col_idx = i; break
                if wh_col_idx is not None: break
        if wh_col_idx is None:
            await status_msg.edit_text("❌ File mein 'warehouse_id' column nahi mila!")
os.unlink(tmp_name); return
        row_start = 2
        all_rows = []
        for row in ws.iter_rows(min_row=row_start, max_row=ws.max_row, values_only=True):
            if any(c is not None and str(c).strip() != '' for c in row):
                all_rows.append(list(row))
        warehouses = sorted(set(str(r[wh_col_idx]).strip() for r in all_rows if r[wh_col_idx] and str(r[wh_col_idx]).strip()))
        if not warehouses:
            await status_msg.edit_text("❌ Koi warehouse data nahi mila!")
            os.unlink(tmp_name); return
        user_id = update.effective_user.id
        user_data[user_id] = {
            'file': tmp_name, 'all_rows': all_rows,
            'headers': headers, 'warehouses': warehouses, 'analysis': None
        }
        keyboard = []
        row = []
        for i, wh in enumerate(warehouses):
            count = sum(1 for r in all_rows if str(r[wh_col_idx]).strip() == wh)
            row.append(InlineKeyboardButton(f"{wh} ({count})", callback_data=f"wh_{i}"))
            if len(row) == 2: keyboard.append(row); row = []
        if row: keyboard.append(row)
        reply_markup = InlineKeyboardMarkup(keyboard)
        await status_msg.edit_text(
            f"✅ File loaded! {len(warehouses)} warehouses mil gaye.\n\n🏭 Warehouse select karo:",
            reply_markup=reply_markup
        )
    except Exception as e:
        logger.error(f"Error: {e}")
        try: os.unlink(tmp_name)
        except: pass
        await status_msg.edit_text(f"❌ Error: {str(e)}")

async def warehouse_callback(update: Update, context):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if user_id not in user_data:
        await query.edit_message_text("❌ Pehle Excel file bhejo!"); return
    wh_index = int(query.data.split('_')[1])
    warehouses = user_data[user_id]['warehouses']
    target_wh = warehouses[wh_index]
    await query.edit_message_text(f"🔍 Analyzing: {target_wh}...")
    try:
        all_rows = user_data[user_id]['all_rows']
        headers = user_data[user_id]['headers']
        a = analyze_warehouse(all_rows, headers, target_wh)
        if a is None:
            await query.edit_message_text("❌ Koi data nahi mila!"); return
        user_data[user_id]['analysis'] = a
        msg = format_analysis(a)
        await query.edit_message_text(msg, parse_mode='Markdown')
    except Exception as e:
        logger.error(f"Analysis error: {e}")
        await query.edit_message_text(f"❌ Analysis error: {str(e)}")

async def report_cmd(update: Update, context):
    user_id = update.effective_user.id
    if user_id not in user_data or not user_data[user_id].get('analysis'):
        await update.message.reply_text("❌ Pehle Excel file bhejo aur analyze karo!"); return
    await update.message.reply_text("📥 Excel report generate ho raha hai...")
    try:
        a = user_data[user_id]['analysis']
        report_fname = generate_report_excel(a)
        with open(report_fname, 'rb') as f:
            await update.message.reply_document(document=f, filename=report_fname, caption=f"📊 Report: {a['wh']}")
        os.unlink(report_fname)
    except Exception as e:
        logger.error(f"Report error: {e}")
        await update.message.reply_text(f"❌ Report error: {str(e)}")

async def all_cmd(update: Update, context):
    user_id = update.effective_user.id
    if user_id not in user_data:
        await update.message.reply_text("❌ Pehle Excel file bhejo!"); return
    try:
        all_rows = user_data[user_id]['all_rows']
        headers = user_data[user_id]['headers']
        warehouses = user_data[user_id]['warehouses']
        msg = "📊 *All Warehouses Summary*\n\n"
        for wh in warehouses:
            a = analyze_warehouse(all_rows, headers, wh)
            if a:
                msg += f"🏭 *{wh}*\n"
                msg += f"  Products: {len(a['prods'])} | Qty: {a['tQ']:,}\n"
                msg += f"  MRP: {fmt(a['tMRP'])} | TF: {fmt(a['tTF'])} | Disc: {a['disc']:.1f}%\n\n"
await update.message.reply_text(msg, parse_mode='Markdown')
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {str(e)}")

def main():
    print("🤖 Warehouse Auction Analyzer Bot v3 starting...")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("report", report_cmd))
    app.add_handler(CommandHandler("all", all_cmd))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(CallbackQueryHandler(warehouse_callback))
    print("✅ Bot v3 is running! Press Ctrl+C to stop.")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()
